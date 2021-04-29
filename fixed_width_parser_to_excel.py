#!/usr/bin/env python3
# fixed_width_parser_to_excel.py

import file_specification as spec
import time
from openpyxl import Workbook, load_workbook

# Define the main program


def main():
    create_spreadsheet(spec.output_file, spec.header, spec.title)
    read_file_write_spreadsheet(
        spec.source_file, spec.output_file, spec.field_lengths)
    input("All done. Press the enter/return key to end.")

# Define function to create spreadsheet and add header


def create_spreadsheet(output_file, header, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for field_name in range(0, len(header)):
        d = ws.cell(row=1, column=field_name+1, value=header[field_name])
    wb.save(output_file)

# Define function to read file and open/write/save spreadsheet


def read_file_write_spreadsheet(source_file, output_file, field_lengths):
    wb = load_workbook(output_file)
    ws = wb.active
    with open(source_file, "r") as file:
        data = file.readlines()
        for data_row in range(0, len(data)):
            for field in range(0, len(field_lengths)-1):
                data_value = data[data_row][field_lengths[field]:field_lengths[field+1]].strip()
                d = ws.cell(row=data_row+2, column=field+1, value=data_value)
    wb.save(output_file)


# Run the main program
main()
